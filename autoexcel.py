import openpyxl as xl
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference

def create_workbook(filename):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Hogwarts Shop"

    sheet.append(["Product", "Original Price"])
    products = [
        ("Wand (Ollivanders)", 150),
        ("Invisibility Cloak", 1200),
        ("Time Turner", 950),
        ("Marauder's Map", 500),
        ("Broomstick (Nimbus 2000)", 850),
        ("Cauldron (Pewter)", 300),
        ("Spell Book", 250),
        ("Chocolate Frogs", 50),
        ("Butterbeer Crate", 100),
        ("Sorting Hat Replica", 400)]

    for product, price in products:
        sheet.append([product, price])

    wb.save(filename)

def apply_discount_and_chart(filename):
    wb = load_workbook(filename)
    sheet = wb["Hogwarts Shop"]

    for row in range(2, sheet.max_row + 1):
        original_price = sheet.cell(row, 2).value
        discounted_price = original_price * 0.9
        sheet.cell(row, 3).value = discounted_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=3, max_col=3)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "E2")

    wb.save(filename),

create_workbook('hogwartsshop.xlsx')
apply_discount_and_chart('hogwartsshop.xlsx')
    


















